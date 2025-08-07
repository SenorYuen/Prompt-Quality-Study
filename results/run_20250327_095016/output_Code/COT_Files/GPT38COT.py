import openpyxl

class ExcelProcessor:
    def __init__(self):
        pass

    def read_excel(self, file_name):
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        return data

    def write_excel(self, data, file_name):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row in data:
                sheet.append(row)
            workbook.save(file_name)
            return 1
        except Exception as e:
            print(f"An error occurred: {e}")
            return 0

    def process_excel_data(self, N, save_file_name):
        data = self.read_excel(save_file_name)
        if not data or N >= len(data[0]):
            return 0, save_file_name

        processed_data = []
        for row in data:
            row = list(row)
            if isinstance(row[N], str):
                row[N] = row[N].upper()
            processed_data.append(tuple(row))

        success = self.write_excel(processed_data, save_file_name)
        return success, save_file_name