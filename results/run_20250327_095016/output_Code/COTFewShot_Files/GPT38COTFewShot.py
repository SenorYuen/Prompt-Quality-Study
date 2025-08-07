import openpyxl

class ExcelProcessor:
    def __init__(self):
        pass

    def read_excel(self, file_name):
        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            return data
        except:
            return None

    def write_excel(self, data, file_name):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row in data:
                sheet.append(row)
            workbook.save(file_name)
            return 1
        except:
            return 0

    def process_excel_data(self, N, save_file_name):
        try:
            data = self.read_excel(save_file_name)
            if data is None:
                return 0, save_file_name
            processed_data = []
            for row in data:
                if len(row) > N:
                    row = list(row)
                    row[N] = str(row[N]).upper()
                processed_data.append(row)
            return self.write_excel(processed_data, save_file_name), save_file_name
        except:
            return 0, save_file_name