import openpyxl


class ExcelProcessor:
    """
    This is a class for processing excel files, including reading and writing excel data,
    as well as processing specific operations and saving as a new excel file.
    """

    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name: str, Excel file name to read
        :return: list of data, Data in Excel
        """
        try:
            # Load the workbook and select the active worksheet
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
            
            # Read data from the sheet
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            
            return data
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            return None

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        """
        try:
            # Create a new workbook and select the active worksheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            
            # Write data to the sheet
            for row in data:
                sheet.append(row)
            
            # Save the workbook to the specified file
            workbook.save(file_name)
            return 1
        except Exception as e:
            print(f"Error writing to Excel file: {e}")
            return 0

    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param save_file_name: str, source file name
        :return: (int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        """
        try:
            # Read data from the specified Excel file
            data = self.read_excel(save_file_name)
            if data is None:
                return 0, save_file_name
            
            # Process the specified column to uppercase
            column_index = N - 1  # Convert to 0-based index
            for i in range(1, len(data)):  # Skip the header row
                row = list(data[i])
                if isinstance(row[column_index], str):
                    row[column_index] = row[column_index].upper()
                data[i] = tuple(row)
            
            # Write the processed data to a new Excel file
            result = self.write_excel(data, save_file_name)
            return result, save_file_name
        except Exception as e:
            print(f"Error processing Excel data: {e}")
            return 0, save_file_name