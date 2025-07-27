import openpyxl

class ExcelProcessor:
    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :return: list of data, Data in Excel
        """
        # Load the workbook and select the active worksheet
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active

        # Initialize a list to store the data
        data = []

        # Iterate through each row in the worksheet and append the row data to the list
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))

        return data

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        """
        try:
            # Create a new workbook and select the active worksheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # Write each row of data to the worksheet
            for row in data:
                sheet.append(row)

            # Save the workbook to the specified file
            workbook.save(file_name)
            return 1  # Return 1 for successful writing
        except Exception as e:
            print(f"An error occurred: {e}")
            return 0  # Return 0 for failed writing

    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :return: (int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        """
        # Read the data from the original Excel file
        data = self.read_excel(save_file_name)

        # Check if the column index N is within the range of available columns
        if len(data) == 0 or N >= len(data[0]):
            raise ValueError("Column index N is out of range")

        # Process the specified column to convert all text to uppercase
        for row in data:
            if isinstance(row[N], str):  # Check if the cell contains a string
                row[N] = row[N].upper()

        # Write the processed data to a new Excel file
        write_result = self.write_excel(data, save_file_name)

        return write_result, save_file_name