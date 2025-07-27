# This is a class for processing excel files, including reading and writing excel data, 
# as well as processing specific operations and saving as a new excel file.

import openpyxl

class ExcelProcessor:
    def __init__(self):
        # Initialize an empty workbook
        self.workbook = None
        self.data = None

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :return: list of data, Data in Excel
        """
        try:
            # Load the workbook from the specified file
            self.workbook = openpyxl.load_workbook(file_name)
            # Select the first sheet
            sheet = self.workbook.active
            # Read data from the sheet
            self.data = [[cell.value for cell in row] for row in sheet.rows]
            return self.data
        except Exception as e:
            # Handle exceptions
            print(f"Error reading Excel file: {e}")
            return None

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        """
        try:
            # Create a new workbook
            self.workbook = openpyxl.Workbook()
            # Select the first sheet
            sheet = self.workbook.active
            # Write data to the sheet
            for i, row in enumerate(data):
                for j, cell in enumerate(row):
                    sheet.cell(row=i+1, column=j+1).value = cell
            # Save the workbook to the specified file
            self.workbook.save(file_name)
            return 1
        except Exception as e:
            # Handle exceptions
            print(f"Error writing Excel file: {e}")
            return 0

    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :return: (int, str), The former is the return value of write_excel, 
                 while the latter is the saved file name of the processed data
        """
        if self.data is None:
            # Check if data has been read
            print("No data to process. Please read an Excel file first.")
            return None
        try:
            # Process the data by converting the specified column to uppercase
            for row in self.data:
                if len(row) > N:
                    row[N] = str(row[N]).upper()
            # Write the processed data to a new Excel file
            result = self.write_excel(self.data, save_file_name)
            return result, save_file_name
        except Exception as e:
            # Handle exceptions
            print(f"Error processing Excel data: {e}")
            return None