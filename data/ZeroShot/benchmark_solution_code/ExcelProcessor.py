'''
# This is a class for processing excel files, including readring and writing excel data, as well as processing specific operations and saving as a new excel file.

import openpyxl


class ExcelProcessor:
    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :return:list of data, Data in Excel
        """


    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        """


    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :return:(int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        """


'''

import openpyxl


class ExcelProcessor:
    def __init__(self):
        pass

    def read_excel(self, file_name):
        data = []
        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            workbook.close()
            return data
        except:
            return None

    def write_excel(self, data, file_name):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row in data:
                sheet.append(row)
            workbook.save(file_name)
            workbook.close()
            return 1
        except:
            return 0

    def process_excel_data(self, N, save_file_name):
        data = self.read_excel(save_file_name)
        if data is None or N >= len(data[0]):
            return 0
        new_data = []
        for row in data:
            new_row = list(row[:])
            if not str(row[N]).isdigit():
                new_row.append(str(row[N]).upper())
            else:
                new_row.append(row[N])
            new_data.append(new_row)
        new_file_name = save_file_name.split('.')[0] + '_process.xlsx'
        success = self.write_excel(new_data, new_file_name)
        return success, new_file_name


